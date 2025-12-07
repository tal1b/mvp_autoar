from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, make_response
from werkzeug.security import generate_password_hash, check_password_hash
import sqlite3
import os
import csv
import io
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

app = Flask(__name__)
app.secret_key = 'your-secret-key-change-in-production'

DATABASE = 'employees.db'

def init_db():
    """Инициализация базы данных"""
    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()
    
    # Таблица пользователей
    c.execute('''CREATE TABLE IF NOT EXISTS users
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  name TEXT NOT NULL,
                  username TEXT UNIQUE NOT NULL,
                  password TEXT NOT NULL,
                  role TEXT NOT NULL DEFAULT 'employee',
                  employee_id INTEGER,
                  FOREIGN KEY (employee_id) REFERENCES employees(id))''')
    
    # Миграция: добавление поля name, если его нет
    try:
        c.execute("ALTER TABLE users ADD COLUMN name TEXT")
        # Обновляем существующие записи
        c.execute("UPDATE users SET name = username WHERE name IS NULL")
    except sqlite3.OperationalError:
        pass  # Колонка уже существует
    
    # Таблица сотрудников (расширенная)
    c.execute('''CREATE TABLE IF NOT EXISTS employees
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  full_name TEXT NOT NULL,
                  position TEXT NOT NULL,
                  department TEXT NOT NULL,
                  start_date TEXT NOT NULL,
                  email TEXT,
                  phone TEXT)''')
    
    # Миграция: добавление полей email и phone, если их нет
    try:
        c.execute("ALTER TABLE employees ADD COLUMN email TEXT")
    except sqlite3.OperationalError:
        pass  # Колонка уже существует
    
    try:
        c.execute("ALTER TABLE employees ADD COLUMN phone TEXT")
    except sqlite3.OperationalError:
        pass  # Колонка уже существует
    
    # Таблица задач
    c.execute('''CREATE TABLE IF NOT EXISTS tasks
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  title TEXT NOT NULL,
                  description TEXT,
                  employee_id INTEGER NOT NULL,
                  status TEXT NOT NULL DEFAULT 'not_started',
                  deadline TEXT,
                  created_at TEXT NOT NULL,
                  FOREIGN KEY (employee_id) REFERENCES employees(id))''')
    
    # Таблица зарплат
    c.execute('''CREATE TABLE IF NOT EXISTS salaries
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  employee_id INTEGER NOT NULL,
                  salary_type TEXT NOT NULL,
                  amount REAL NOT NULL,
                  hourly_rate REAL,
                  period_start TEXT NOT NULL,
                  period_end TEXT NOT NULL,
                  created_at TEXT NOT NULL,
                  FOREIGN KEY (employee_id) REFERENCES employees(id))''')
    
    # Таблица смен
    c.execute('''CREATE TABLE IF NOT EXISTS shifts
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  employee_id INTEGER NOT NULL,
                  shift_date TEXT NOT NULL,
                  start_time TEXT NOT NULL,
                  end_time TEXT NOT NULL,
                  hours REAL,
                  FOREIGN KEY (employee_id) REFERENCES employees(id))''')
    
    # Создаем тестового менеджера по умолчанию (только если нет пользователей)
    c.execute("SELECT COUNT(*) FROM users")
    if c.fetchone()[0] == 0:
        manager_password = generate_password_hash('manager123')
        c.execute("INSERT INTO users (name, username, password, role) VALUES (?, ?, ?, ?)",
                  ('Test Manager', 'manager', manager_password, 'manager'))
        
        # Создаем тестового сотрудника
        c.execute('''INSERT INTO employees (full_name, position, department, start_date, email, phone)
                     VALUES (?, ?, ?, ?, ?, ?)''',
                  ('Test Employee', 'Developer', 'IT', '2024-01-01', 'test@example.com', '+1 (555) 123-4567'))
        employee_id = c.lastrowid
        
        # Создаем пользователя для сотрудника
        employee_password = generate_password_hash('employee123')
        c.execute("INSERT INTO users (name, username, password, role, employee_id) VALUES (?, ?, ?, ?, ?)",
                  ('Test Employee', 'employee', employee_password, 'employee', employee_id))
    
    conn.commit()
    conn.close()

def get_db():
    """Получение соединения с базой данных"""
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def validate_required_fields(data, required_fields):
    """Валидация обязательных полей"""
    missing = [field for field in required_fields if not data.get(field)]
    return missing

def is_manager():
    """Проверка, является ли пользователь менеджером"""
    return session.get('role') == 'manager'

def is_employee():
    """Проверка, является ли пользователь сотрудником"""
    return session.get('role') == 'employee'

def get_employee_id():
    """Получение ID сотрудника из сессии"""
    if is_employee():
        conn = get_db()
        user = conn.execute('SELECT employee_id FROM users WHERE id = ?', (session['user_id'],)).fetchone()
        conn.close()
        return user['employee_id'] if user else None
    return None

# ==================== АУТЕНТИФИКАЦИЯ ====================

@app.route('/')
def index():
    """Главная страница"""
    if 'user_id' in session:
        if is_manager():
            return redirect(url_for('dashboard'))
        else:
            return redirect(url_for('my_tasks'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Страница входа"""
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        conn = get_db()
        user = conn.execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()
        conn.close()
        
        if user and check_password_hash(user['password'], password):
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['name'] = user['name']
            session['role'] = user['role']
            session['employee_id'] = user['employee_id']
            flash('You have successfully logged in', 'success')
            if is_manager():
                return redirect(url_for('dashboard'))
            else:
                return redirect(url_for('my_tasks'))
        else:
            flash('Invalid username or password', 'error')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    """Выход из системы"""
    session.clear()
    flash('You have logged out', 'info')
    return redirect(url_for('login'))

# ==================== ГЛАВНАЯ СТРАНИЦА ====================

@app.route('/register', methods=['GET', 'POST'])
def register():
    """Страница регистрации"""
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        role = request.form.get('role', '').strip()
        
        if not all([name, username, password, role]):
            flash('All fields are required', 'error')
            return render_template('register.html')
        
        if role not in ['manager', 'employee']:
            flash('Invalid role', 'error')
            return render_template('register.html')
        
        conn = get_db()
        # Проверяем, существует ли пользователь с таким логином
        existing = conn.execute('SELECT id FROM users WHERE username = ?', (username,)).fetchone()
        if existing:
            conn.close()
            flash('User with this username already exists', 'error')
            return render_template('register.html')
        
        # Если регистрируется сотрудник, нужно создать запись в employees
        employee_id = None
        if role == 'employee':
            # Создаем запись сотрудника
            cursor = conn.cursor()
            cursor.execute('''INSERT INTO employees (full_name, position, department, start_date)
                           VALUES (?, ?, ?, ?)''',
                        (name, 'Not specified', 'Not specified', datetime.now().strftime('%Y-%m-%d')))
            employee_id = cursor.lastrowid
        
        # Создаем пользователя
        password_hash = generate_password_hash(password)
        cursor = conn.cursor()
        cursor.execute('''INSERT INTO users (name, username, password, role, employee_id)
                       VALUES (?, ?, ?, ?, ?)''',
                    (name, username, password_hash, role, employee_id))
        conn.commit()
        conn.close()
        
        flash('Registration successful! Please log in', 'success')
        return redirect(url_for('login'))
    
    return render_template('register.html')

@app.route('/dashboard')
def dashboard():
    """Главная страница для менеджера"""
    if 'user_id' not in session or not is_manager():
        return redirect(url_for('login'))
    
    conn = get_db()
    
    # Статистика
    stats = {
        'total_employees': conn.execute('SELECT COUNT(*) FROM employees').fetchone()[0],
        'total_tasks': conn.execute('SELECT COUNT(*) FROM tasks').fetchone()[0],
        'tasks_by_status': {},
        'employees_by_department': {},
        'total_salaries': 0
    }
    
    # Задачи по статусу
    for status in ['not_started', 'in_progress', 'completed']:
        count = conn.execute('SELECT COUNT(*) FROM tasks WHERE status = ?', (status,)).fetchone()[0]
        stats['tasks_by_status'][status] = count
    
    # Сотрудники по отделам
    depts = conn.execute('SELECT department, COUNT(*) as count FROM employees GROUP BY department').fetchall()
    stats['employees_by_department'] = {row['department']: row['count'] for row in depts}
    
    # Общая сумма зарплат за текущий месяц
    current_month = datetime.now().strftime('%Y-%m')
    salaries = conn.execute('''SELECT SUM(amount) as total FROM salaries 
                               WHERE period_start LIKE ?''', (f'{current_month}%',)).fetchone()
    stats['total_salaries'] = salaries['total'] or 0
    
    conn.close()
    
    return render_template('dashboard.html', stats=stats)

# ==================== СОТРУДНИКИ ====================

@app.route('/employees')
def employees():
    """Страница со списком сотрудников (только просмотр)"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Только менеджер может видеть список всех сотрудников
    if not is_manager():
        flash('Access denied', 'error')
        return redirect(url_for('my_tasks'))
    
    search = request.args.get('search', '')
    department_filter = request.args.get('department', '')
    
    conn = get_db()
    query = 'SELECT * FROM employees WHERE 1=1'
    params = []
    
    if search:
        query += ' AND (full_name LIKE ? OR position LIKE ?)'
        params.extend([f'%{search}%', f'%{search}%'])
    
    if department_filter:
        query += ' AND department = ?'
        params.append(department_filter)
    
    query += ' ORDER BY id DESC'
    
    employees_list = conn.execute(query, params).fetchall()
    
    # Получаем роли для каждого сотрудника
    employees_with_roles = []
    for emp in employees_list:
        user = conn.execute('SELECT role FROM users WHERE employee_id = ?', (emp['id'],)).fetchone()
        emp_dict = dict(emp)
        emp_dict['role'] = user['role'] if user else None
        employees_with_roles.append(emp_dict)
    
    # Получаем список отделов для фильтра
    departments = conn.execute('SELECT DISTINCT department FROM employees ORDER BY department').fetchall()
    
    conn.close()
    
    return render_template('employees.html', 
                         employees=employees_with_roles,
                         departments=departments,
                         search=search,
                         department_filter=department_filter,
                         is_manager=is_manager())

@app.route('/employee/add', methods=['GET', 'POST'])
def add_employee():
    """Добавление сотрудника - отключено (только просмотр)"""
    flash('Adding employees is not available. Use registration', 'error')
    return redirect(url_for('employees'))
    
    if request.method == 'POST':
        data = {
            'full_name': request.form.get('full_name', '').strip(),
            'position': request.form.get('position', '').strip(),
            'department': request.form.get('department', '').strip(),
            'start_date': request.form.get('start_date', '').strip(),
            'email': request.form.get('email', '').strip(),
            'phone': request.form.get('phone', '').strip()
        }
        
        required = ['full_name', 'position', 'department', 'start_date']
        missing = validate_required_fields(data, required)
        
        if missing:
            flash(f'Required fields are missing: {", ".join(missing)}', 'error')
            return render_template('employee_form.html', employee=None)
        
        # Валидация email
        if data['email'] and '@' not in data['email']:
            flash('Invalid email address', 'error')
            return render_template('employee_form.html', employee=None)
        
        conn = get_db()
        conn.execute('''INSERT INTO employees (full_name, position, department, start_date, email, phone)
                       VALUES (?, ?, ?, ?, ?, ?)''',
                    (data['full_name'], data['position'], data['department'], 
                     data['start_date'], data['email'] or None, data['phone'] or None))
        conn.commit()
        conn.close()
        
        flash('Employee added successfully', 'success')
        return redirect(url_for('employees'))
    
    return render_template('employee_form.html', employee=None)

@app.route('/my-profile', methods=['GET', 'POST'])
def my_profile():
    """Редактирование своего профиля (для сотрудника)"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    employee_id = get_employee_id()
    if not employee_id:
        flash('Employee not found', 'error')
        return redirect(url_for('my_tasks'))
    
    conn = get_db()
    
    if request.method == 'POST':
        data = {
            'full_name': request.form.get('full_name', '').strip(),
            'position': request.form.get('position', '').strip(),
            'department': request.form.get('department', '').strip(),
            'email': request.form.get('email', '').strip(),
            'phone': request.form.get('phone', '').strip()
        }
        
        required = ['full_name', 'position', 'department']
        missing = validate_required_fields(data, required)
        
        if missing:
            flash(f'Required fields are missing: {", ".join(missing)}', 'error')
            employee = conn.execute('SELECT * FROM employees WHERE id = ?', (employee_id,)).fetchone()
            conn.close()
            return render_template('my_profile.html', employee=employee)
        
        # Валидация email
        if data['email'] and '@' not in data['email']:
            flash('Invalid email address', 'error')
            employee = conn.execute('SELECT * FROM employees WHERE id = ?', (employee_id,)).fetchone()
            conn.close()
            return render_template('my_profile.html', employee=employee)
        
        # Обновляем данные сотрудника (без start_date и role)
        conn.execute('''UPDATE employees 
                       SET full_name = ?, position = ?, department = ?, 
                           email = ?, phone = ?
                       WHERE id = ?''',
                    (data['full_name'], data['position'], data['department'], 
                     data['email'] or None, data['phone'] or None, employee_id))
        
        conn.commit()
        conn.close()
        
        flash('Profile updated successfully', 'success')
        return redirect(url_for('my_profile'))
    
    employee = conn.execute('SELECT * FROM employees WHERE id = ?', (employee_id,)).fetchone()
    conn.close()
    
    if not employee:
        flash('Employee not found', 'error')
        return redirect(url_for('my_tasks'))
    
    return render_template('my_profile.html', employee=employee)

@app.route('/employee/edit/<int:employee_id>', methods=['GET', 'POST'])
def edit_employee(employee_id):
    """Редактирование сотрудника (только для менеджера)"""
    if not is_manager():
        flash('Access denied', 'error')
        return redirect(url_for('employees'))
    
    conn = get_db()
    
    if request.method == 'POST':
        data = {
            'full_name': request.form.get('full_name', '').strip(),
            'position': request.form.get('position', '').strip(),
            'department': request.form.get('department', '').strip(),
            'start_date': request.form.get('start_date', '').strip(),
            'email': request.form.get('email', '').strip(),
            'phone': request.form.get('phone', '').strip(),
            'role': request.form.get('role', '').strip()
        }
        
        required = ['full_name', 'position', 'department', 'start_date']
        missing = validate_required_fields(data, required)
        
        if missing:
            flash(f'Required fields are missing: {", ".join(missing)}', 'error')
            employee = conn.execute('SELECT * FROM employees WHERE id = ?', (employee_id,)).fetchone()
            # Получаем текущую роль пользователя
            user = conn.execute('SELECT role FROM users WHERE employee_id = ?', (employee_id,)).fetchone()
            conn.close()
            return render_template('employee_form.html', employee=employee, current_role=user['role'] if user else None)
        
        # Валидация email
        if data['email'] and '@' not in data['email']:
            flash('Invalid email address', 'error')
            employee = conn.execute('SELECT * FROM employees WHERE id = ?', (employee_id,)).fetchone()
            user = conn.execute('SELECT role FROM users WHERE employee_id = ?', (employee_id,)).fetchone()
            conn.close()
            return render_template('employee_form.html', employee=employee, current_role=user['role'] if user else None)
        
        # Обновляем данные сотрудника
        conn.execute('''UPDATE employees 
                       SET full_name = ?, position = ?, department = ?, start_date = ?, 
                           email = ?, phone = ?
                       WHERE id = ?''',
                    (data['full_name'], data['position'], data['department'], 
                     data['start_date'], data['email'] or None, data['phone'] or None, employee_id))
        
        # Обновляем роль пользователя, если она указана
        if data['role'] and data['role'] in ['manager', 'employee']:
            user = conn.execute('SELECT id FROM users WHERE employee_id = ?', (employee_id,)).fetchone()
            if user:
                conn.execute('UPDATE users SET role = ? WHERE employee_id = ?', 
                           (data['role'], employee_id))
        
        conn.commit()
        conn.close()
        
        flash('Employee updated successfully', 'success')
        return redirect(url_for('employees'))
    
    employee = conn.execute('SELECT * FROM employees WHERE id = ?', (employee_id,)).fetchone()
    # Получаем текущую роль пользователя
    user = conn.execute('SELECT role FROM users WHERE employee_id = ?', (employee_id,)).fetchone()
    conn.close()
    
    if not employee:
        flash('Employee not found', 'error')
        return redirect(url_for('employees'))
    
    return render_template('employee_form.html', employee=employee, current_role=user['role'] if user else None)

@app.route('/employee/delete/<int:employee_id>', methods=['POST'])
def delete_employee(employee_id):
    """Удаление сотрудника (только для менеджера)"""
    if not is_manager():
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    
    conn = get_db()
    
    # Проверяем, существует ли сотрудник
    employee = conn.execute('SELECT * FROM employees WHERE id = ?', (employee_id,)).fetchone()
    if not employee:
        conn.close()
        return jsonify({'success': False, 'message': 'Employee not found'}), 404
    
    # Проверяем, не пытается ли менеджер удалить самого себя
    current_user = conn.execute('SELECT employee_id FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    if current_user and current_user['employee_id'] == employee_id:
        conn.close()
        return jsonify({'success': False, 'message': 'You cannot delete yourself'}), 400
    
    # Удаляем связанные данные (задачи, зарплаты, смены)
    conn.execute('DELETE FROM tasks WHERE employee_id = ?', (employee_id,))
    conn.execute('DELETE FROM salaries WHERE employee_id = ?', (employee_id,))
    conn.execute('DELETE FROM shifts WHERE employee_id = ?', (employee_id,))
    
    # Удаляем пользователя, связанного с сотрудником
    conn.execute('DELETE FROM users WHERE employee_id = ?', (employee_id,))
    
    # Удаляем сотрудника
    conn.execute('DELETE FROM employees WHERE id = ?', (employee_id,))
    
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'message': 'Employee and all related data deleted successfully'})

# ==================== ЗАДАЧИ ====================

@app.route('/tasks')
def tasks():
    """Список задач (для админа - все, для сотрудника - только свои)"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    status_filter = request.args.get('status', '')
    employee_filter = request.args.get('employee', '')
    search = request.args.get('search', '')
    
    conn = get_db()
    
    if is_manager():
        query = '''SELECT t.*, e.full_name as employee_name 
                   FROM tasks t 
                   JOIN employees e ON t.employee_id = e.id 
                   WHERE 1=1'''
        params = []
    else:
        employee_id = get_employee_id()
        if not employee_id:
            conn.close()
            flash('Error: employee not found', 'error')
            return redirect(url_for('login'))
        query = '''SELECT t.*, e.full_name as employee_name 
                   FROM tasks t 
                   JOIN employees e ON t.employee_id = e.id 
                   WHERE t.employee_id = ?'''
        params = [employee_id]
    
    if status_filter:
        query += ' AND t.status = ?'
        params.append(status_filter)
    
    if employee_filter and is_manager():
        query += ' AND t.employee_id = ?'
        params.append(employee_filter)
    
    if search:
        query += ' AND (t.title LIKE ? OR t.description LIKE ?)'
        params.extend([f'%{search}%', f'%{search}%'])
    
    query += ' ORDER BY t.created_at DESC'
    
    tasks_list = conn.execute(query, params).fetchall()
    
    # Получаем список сотрудников для фильтра (только для админа)
    employees_list = []
    if is_manager():
        employees_list = conn.execute('SELECT id, full_name FROM employees ORDER BY full_name').fetchall()
    
    conn.close()
    
    return render_template('tasks.html',
                         tasks=tasks_list,
                         employees=employees_list,
                         status_filter=status_filter,
                         employee_filter=employee_filter,
                         search=search,
                         is_manager=is_manager())

@app.route('/my-tasks')
def my_tasks():
    """Мои задачи (для сотрудника)"""
    if not is_employee():
        return redirect(url_for('tasks'))
    return redirect(url_for('tasks'))

@app.route('/task/add', methods=['GET', 'POST'])
def add_task():
    """Добавление задачи (только для админа)"""
    if not is_manager():
        flash('Access denied', 'error')
        return redirect(url_for('tasks'))
    
    if request.method == 'POST':
        data = {
            'title': request.form.get('title', '').strip(),
            'description': request.form.get('description', '').strip(),
            'employee_id': request.form.get('employee_id', '').strip(),
            'status': request.form.get('status', 'not_started').strip(),
            'deadline': request.form.get('deadline', '').strip()
        }
        
        required = ['title', 'employee_id']
        missing = validate_required_fields(data, required)
        
        if missing:
            flash(f'Required fields are missing: {", ".join(missing)}', 'error')
            conn = get_db()
            employees_list = conn.execute('SELECT id, full_name FROM employees ORDER BY full_name').fetchall()
            conn.close()
            return render_template('task_form.html', task=None, employees=employees_list, is_manager=True)
        
        conn = get_db()
        conn.execute('''INSERT INTO tasks (title, description, employee_id, status, deadline, created_at)
                       VALUES (?, ?, ?, ?, ?, ?)''',
                    (data['title'], data['description'] or None, int(data['employee_id']), 
                     data['status'], data['deadline'] or None, datetime.now().isoformat()))
        conn.commit()
        conn.close()
        
        flash('Task added successfully', 'success')
        return redirect(url_for('tasks'))
    
    conn = get_db()
    employees_list = conn.execute('SELECT id, full_name FROM employees ORDER BY full_name').fetchall()
    conn.close()
    
    if not employees_list:
        flash('No available employees. Please register employees first.', 'error')
        return redirect(url_for('employees'))
    
    return render_template('task_form.html', task=None, employees=employees_list, is_manager=True)

@app.route('/task/edit/<int:task_id>', methods=['GET', 'POST'])
def edit_task(task_id):
    """Редактирование задачи"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    conn = get_db()
    task = conn.execute('SELECT * FROM tasks WHERE id = ?', (task_id,)).fetchone()
    
    if not task:
        conn.close()
        flash('Task not found', 'error')
        return redirect(url_for('tasks'))
    
    # Проверка прав доступа
    if is_employee() and task['employee_id'] != get_employee_id():
        conn.close()
        flash('Access denied', 'error')
        return redirect(url_for('tasks'))
    
    if request.method == 'POST':
        data = {
            'title': request.form.get('title', '').strip(),
            'description': request.form.get('description', '').strip(),
            'status': request.form.get('status', 'not_started').strip(),
            'deadline': request.form.get('deadline', '').strip()
        }
        
        if is_manager():
            data['employee_id'] = request.form.get('employee_id', '').strip()
        
        required = ['title']
        if is_manager():
            required.append('employee_id')
        
        missing = validate_required_fields(data, required)
        
        if missing:
            flash(f'Required fields are missing: {", ".join(missing)}', 'error')
            employees_list = []
            if is_manager():
                employees_list = conn.execute('SELECT id, full_name FROM employees ORDER BY full_name').fetchall()
            conn.close()
            return render_template('task_form.html', task=task, employees=employees_list, is_admin=is_manager())
        
        if is_manager():
            conn.execute('''UPDATE tasks 
                           SET title = ?, description = ?, employee_id = ?, status = ?, deadline = ?
                           WHERE id = ?''',
                        (data['title'], data['description'] or None, int(data['employee_id']),
                         data['status'], data['deadline'] or None, task_id))
        else:
            conn.execute('''UPDATE tasks 
                           SET title = ?, description = ?, status = ?, deadline = ?
                           WHERE id = ?''',
                        (data['title'], data['description'] or None, data['status'],
                         data['deadline'] or None, task_id))
        
        conn.commit()
        conn.close()
        
        flash('Task updated successfully', 'success')
        return redirect(url_for('tasks'))
    
    employees_list = []
    if is_manager():
        employees_list = conn.execute('SELECT id, full_name FROM employees ORDER BY full_name').fetchall()
    
    conn.close()
    
    return render_template('task_form.html', task=task, employees=employees_list, is_admin=is_manager())

@app.route('/task/delete/<int:task_id>', methods=['POST'])
def delete_task(task_id):
    """Удаление задачи (только для админа)"""
    if not is_manager():
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    
    conn = get_db()
    conn.execute('DELETE FROM tasks WHERE id = ?', (task_id,))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'message': 'Task deleted successfully'})

# ==================== ЗАРПЛАТЫ ====================

@app.route('/salaries')
def salaries():
    """Список зарплат"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    employee_filter = request.args.get('employee', '')
    period_filter = request.args.get('period', '')
    
    conn = get_db()
    
    if is_manager():
        query = '''SELECT s.*, e.full_name as employee_name 
                   FROM salaries s 
                   JOIN employees e ON s.employee_id = e.id 
                   WHERE 1=1'''
        params = []
    else:
        employee_id = get_employee_id()
        if not employee_id:
            conn.close()
            flash('Error: employee not found', 'error')
            return redirect(url_for('login'))
        query = '''SELECT s.*, e.full_name as employee_name 
                   FROM salaries s 
                   JOIN employees e ON s.employee_id = e.id 
                   WHERE s.employee_id = ?'''
        params = [employee_id]
    
    if employee_filter and is_manager():
        query += ' AND s.employee_id = ?'
        params.append(employee_filter)
    
    if period_filter:
        query += ' AND s.period_start LIKE ?'
        params.append(f'{period_filter}%')
    
    query += ' ORDER BY s.period_start DESC, s.created_at DESC'
    
    salaries_list = conn.execute(query, params).fetchall()
    
    employees_list = []
    if is_manager():
        employees_list = conn.execute('SELECT id, full_name FROM employees ORDER BY full_name').fetchall()
    
    conn.close()
    
    return render_template('salaries.html',
                         salaries=salaries_list,
                         employees=employees_list,
                         employee_filter=employee_filter,
                         period_filter=period_filter,
                         is_manager=is_manager())

@app.route('/salary/add', methods=['GET', 'POST'])
def add_salary():
    """Добавление зарплаты (только для админа)"""
    if not is_manager():
        flash('Access denied', 'error')
        return redirect(url_for('salaries'))
    
    if request.method == 'POST':
        data = {
            'employee_id': request.form.get('employee_id', '').strip(),
            'salary_type': request.form.get('salary_type', '').strip(),
            'amount': request.form.get('amount', '').strip(),
            'hourly_rate': request.form.get('hourly_rate', '').strip(),
            'period_start': request.form.get('period_start', '').strip(),
            'period_end': request.form.get('period_end', '').strip()
        }
        
        required = ['employee_id', 'salary_type', 'period_start', 'period_end']
        missing = validate_required_fields(data, required)
        
        if missing:
            flash(f'Required fields are missing: {", ".join(missing)}', 'error')
            conn = get_db()
            employees_list = conn.execute('SELECT id, full_name FROM employees ORDER BY full_name').fetchall()
            conn.close()
            return render_template('salary_form.html', salary=None, employees=employees_list)
        
        # Валидация типа зарплаты
        if data['salary_type'] == 'fixed' and not data['amount']:
            flash('For fixed salary, amount is required', 'error')
            conn = get_db()
            employees_list = conn.execute('SELECT id, full_name FROM employees ORDER BY full_name').fetchall()
            conn.close()
            return render_template('salary_form.html', salary=None, employees=employees_list)
        
        if data['salary_type'] == 'hourly' and not data['hourly_rate']:
            flash('For hourly salary, rate is required', 'error')
            conn = get_db()
            employees_list = conn.execute('SELECT id, full_name FROM employees ORDER BY full_name').fetchall()
            conn.close()
            return render_template('salary_form.html', salary=None, employees=employees_list)
        
        try:
            amount = float(data['amount']) if data['amount'] else 0
            hourly_rate = float(data['hourly_rate']) if data['hourly_rate'] else None
        except ValueError:
            flash('Invalid amount or rate value', 'error')
            conn = get_db()
            employees_list = conn.execute('SELECT id, full_name FROM employees ORDER BY full_name').fetchall()
            conn.close()
            return render_template('salary_form.html', salary=None, employees=employees_list)
        
        conn = get_db()
        conn.execute('''INSERT INTO salaries (employee_id, salary_type, amount, hourly_rate, 
                                             period_start, period_end, created_at)
                       VALUES (?, ?, ?, ?, ?, ?, ?)''',
                    (int(data['employee_id']), data['salary_type'], amount, hourly_rate,
                     data['period_start'], data['period_end'], datetime.now().isoformat()))
        conn.commit()
        conn.close()
        
        flash('Salary added successfully', 'success')
        return redirect(url_for('salaries'))
    
    conn = get_db()
    employees_list = conn.execute('SELECT id, full_name FROM employees ORDER BY full_name').fetchall()
    conn.close()
    
    return render_template('salary_form.html', salary=None, employees=employees_list)

# ==================== СМЕНЫ ====================

@app.route('/shifts')
def shifts():
    """Список смен"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    employee_filter = request.args.get('employee', '')
    date_filter = request.args.get('date', '')
    
    conn = get_db()
    
    if is_manager():
        query = '''SELECT s.*, e.full_name as employee_name 
                   FROM shifts s 
                   JOIN employees e ON s.employee_id = e.id 
                   WHERE 1=1'''
        params = []
    else:
        employee_id = get_employee_id()
        if not employee_id:
            conn.close()
            flash('Error: employee not found', 'error')
            return redirect(url_for('login'))
        query = '''SELECT s.*, e.full_name as employee_name 
                   FROM shifts s 
                   JOIN employees e ON s.employee_id = e.id 
                   WHERE s.employee_id = ?'''
        params = [employee_id]
    
    if employee_filter and is_manager():
        query += ' AND s.employee_id = ?'
        params.append(employee_filter)
    
    if date_filter:
        query += ' AND s.shift_date = ?'
        params.append(date_filter)
    
    query += ' ORDER BY s.shift_date DESC, s.start_time'
    
    shifts_list = conn.execute(query, params).fetchall()
    
    employees_list = []
    if is_manager():
        employees_list = conn.execute('SELECT id, full_name FROM employees ORDER BY full_name').fetchall()
    
    conn.close()
    
    return render_template('shifts.html',
                         shifts=shifts_list,
                         employees=employees_list,
                         employee_filter=employee_filter,
                         date_filter=date_filter,
                         is_manager=is_manager())

@app.route('/shift/add', methods=['GET', 'POST'])
def add_shift():
    """Добавление смены (только для админа)"""
    if not is_manager():
        flash('Access denied', 'error')
        return redirect(url_for('shifts'))
    
    if request.method == 'POST':
        data = {
            'employee_id': request.form.get('employee_id', '').strip(),
            'shift_date': request.form.get('shift_date', '').strip(),
            'start_time': request.form.get('start_time', '').strip(),
            'end_time': request.form.get('end_time', '').strip()
        }
        
        required = ['employee_id', 'shift_date', 'start_time', 'end_time']
        missing = validate_required_fields(data, required)
        
        if missing:
            flash(f'Required fields are missing: {", ".join(missing)}', 'error')
            conn = get_db()
            employees_list = conn.execute('SELECT id, full_name FROM employees ORDER BY full_name').fetchall()
            conn.close()
            return render_template('shift_form.html', shift=None, employees=employees_list)
        
        # Вычисление часов
        try:
            start = datetime.strptime(f"{data['shift_date']} {data['start_time']}", "%Y-%m-%d %H:%M")
            end = datetime.strptime(f"{data['shift_date']} {data['end_time']}", "%Y-%m-%d %H:%M")
            if end < start:
                end = datetime.strptime(f"{data['shift_date']} {data['end_time']}", "%Y-%m-%d %H:%M").replace(day=end.day + 1)
            hours = (end - start).total_seconds() / 3600
        except ValueError:
            flash('Invalid time format', 'error')
            conn = get_db()
            employees_list = conn.execute('SELECT id, full_name FROM employees ORDER BY full_name').fetchall()
            conn.close()
            return render_template('shift_form.html', shift=None, employees=employees_list)
        
        conn = get_db()
        conn.execute('''INSERT INTO shifts (employee_id, shift_date, start_time, end_time, hours)
                       VALUES (?, ?, ?, ?, ?)''',
                    (int(data['employee_id']), data['shift_date'], data['start_time'], 
                     data['end_time'], hours))
        conn.commit()
        conn.close()
        
        flash('Shift added successfully', 'success')
        return redirect(url_for('shifts'))
    
    conn = get_db()
    employees_list = conn.execute('SELECT id, full_name FROM employees ORDER BY full_name').fetchall()
    conn.close()
    
    return render_template('shift_form.html', shift=None, employees=employees_list)

@app.route('/shift/delete/<int:shift_id>', methods=['POST'])
def delete_shift(shift_id):
    """Удаление смены (только для админа)"""
    if not is_manager():
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    
    conn = get_db()
    conn.execute('DELETE FROM shifts WHERE id = ?', (shift_id,))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'message': 'Shift deleted successfully'})

# ==================== ЭКСПОРТ ДАННЫХ ====================

@app.route('/export/<data_type>')
def export_data(data_type):
    """Экспорт данных в CSV или Excel"""
    if not is_manager():
        flash('Access denied', 'error')
        return redirect(url_for('dashboard'))
    
    format_type = request.args.get('format', 'csv')
    
    conn = get_db()
    
    if data_type == 'employees':
        data = conn.execute('SELECT * FROM employees ORDER BY id').fetchall()
        filename = 'employees'
        headers = ['ID', 'Full Name', 'Position', 'Department', 'Start Date', 'Email', 'Phone']
        
    elif data_type == 'tasks':
        data = conn.execute('''SELECT t.id, t.title, t.description, e.full_name as employee, 
                               t.status, t.deadline, t.created_at 
                               FROM tasks t 
                               JOIN employees e ON t.employee_id = e.id 
                               ORDER BY t.id''').fetchall()
        filename = 'tasks'
        headers = ['ID', 'Title', 'Description', 'Employee', 'Status', 'Deadline', 'Created']
        
    elif data_type == 'salaries':
        data = conn.execute('''SELECT s.id, e.full_name as employee, s.salary_type, 
                               s.amount, s.hourly_rate, s.period_start, s.period_end, s.created_at 
                               FROM salaries s 
                               JOIN employees e ON s.employee_id = e.id 
                               ORDER BY s.id''').fetchall()
        filename = 'salaries'
        headers = ['ID', 'Employee', 'Type', 'Amount', 'Hourly Rate', 'Period Start', 'Period End', 'Created']
        
    elif data_type == 'shifts':
        data = conn.execute('''SELECT s.id, e.full_name as employee, s.shift_date, 
                               s.start_time, s.end_time, s.hours 
                               FROM shifts s 
                               JOIN employees e ON s.employee_id = e.id 
                               ORDER BY s.id''').fetchall()
        filename = 'shifts'
        headers = ['ID', 'Employee', 'Date', 'Start', 'End', 'Hours']
        
    else:
        conn.close()
        flash('Unknown data type', 'error')
        return redirect(url_for('dashboard'))
    
    conn.close()
    
    if format_type == 'excel':
        return export_excel(data, headers, filename)
    else:
        return export_csv(data, headers, filename)

def export_csv(data, headers, filename):
    """Экспорт в CSV"""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    
    for row in data:
        writer.writerow([row[i] for i in range(len(row))])
    
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}.csv'
    return response

def export_excel(data, headers, filename):
    """Экспорт в Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = filename
    
    # Заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Данные
    for row_idx, row in enumerate(data, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Автоширина колонок
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}.xlsx'
    return response

if __name__ == '__main__':
    init_db()
    app.run(debug=True, host='0.0.0.0', port=5000)
